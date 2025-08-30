import openpyxl
from openpyxl import Workbook
import os
import logging
from datetime import datetime

class ExcelManager:
    def __init__(self):
        self.users_file = 'users.xlsx'
        self.admins_file = 'admins.xlsx'
        self.init_excel_files()
    
    def init_excel_files(self):
        """Initialize Excel files if they don't exist"""
        # Initialize users file
        if not os.path.exists(self.users_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(['ID', 'Username', 'Email', 'Password', 'Created Date'])
            wb.save(self.users_file)
            logging.info("Created users.xlsx file")
        
        # Initialize admins file
        if not os.path.exists(self.admins_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Admins"
            ws.append(['ID', 'Username', 'Email', 'Password', 'Admin Name', 'Created Date'])
            wb.save(self.admins_file)
            logging.info("Created admins.xlsx file")
    
    def get_next_id(self, filename):
        """Get the next available ID for a file"""
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
            
            return max_id + 1
        except Exception as e:
            logging.error(f"Error getting next ID: {str(e)}")
            return 1
    
    def add_user(self, user_data):
        """Add a new user to the Excel file"""
        try:
            wb = openpyxl.load_workbook(self.users_file)
            ws = wb.active
            
            user_id = self.get_next_id(self.users_file)
            created_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            ws.append([
                user_id,
                user_data['username'],
                user_data['email'],
                user_data['password'],
                created_date
            ])
            
            wb.save(self.users_file)
            logging.info(f"Added user: {user_data['username']}")
            return user_id
        except Exception as e:
            logging.error(f"Error adding user: {str(e)}")
            raise
    
    def add_admin(self, admin_data):
        """Add a new admin to the Excel file"""
        try:
            wb = openpyxl.load_workbook(self.admins_file)
            ws = wb.active
            
            admin_id = self.get_next_id(self.admins_file)
            created_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            ws.append([
                admin_id,
                admin_data['username'],
                admin_data['email'],
                admin_data['password'],
                admin_data['admin_name'],
                created_date
            ])
            
            wb.save(self.admins_file)
            logging.info(f"Added admin: {admin_data['username']}")
            return admin_id
        except Exception as e:
            logging.error(f"Error adding admin: {str(e)}")
            raise
    
    def get_user(self, username):
        """Get user by username"""
        try:
            wb = openpyxl.load_workbook(self.users_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == username:  # Username is in column 2 (index 1)
                    return {
                        'id': row[0],
                        'username': row[1],
                        'email': row[2],
                        'password': row[3],
                        'created_date': row[4]
                    }
            return None
        except Exception as e:
            logging.error(f"Error getting user: {str(e)}")
            return None
    
    def get_admin(self, username):
        """Get admin by username"""
        try:
            wb = openpyxl.load_workbook(self.admins_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == username:  # Username is in column 2 (index 1)
                    return {
                        'id': row[0],
                        'username': row[1],
                        'email': row[2],
                        'password': row[3],
                        'admin_name': row[4],
                        'created_date': row[5]
                    }
            return None
        except Exception as e:
            logging.error(f"Error getting admin: {str(e)}")
            return None
    
    def get_all_users(self):
        """Get all users"""
        try:
            wb = openpyxl.load_workbook(self.users_file)
            ws = wb.active
            
            users = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If ID exists
                    users.append({
                        'id': row[0],
                        'username': row[1],
                        'email': row[2],
                        'created_date': row[4]
                    })
            return users
        except Exception as e:
            logging.error(f"Error getting all users: {str(e)}")
            return []
    
    def get_all_admins(self):
        """Get all admins"""
        try:
            wb = openpyxl.load_workbook(self.admins_file)
            ws = wb.active
            
            admins = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If ID exists
                    admins.append({
                        'id': row[0],
                        'username': row[1],
                        'email': row[2],
                        'admin_name': row[4],
                        'created_date': row[5]
                    })
            return admins
        except Exception as e:
            logging.error(f"Error getting all admins: {str(e)}")
            return []
    
    def delete_user(self, user_id):
        """Delete user by ID"""
        try:
            wb = openpyxl.load_workbook(self.users_file)
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == user_id:
                    ws.delete_rows(row_num)
                    wb.save(self.users_file)
                    logging.info(f"Deleted user ID: {user_id}")
                    return True
            
            logging.warning(f"User ID {user_id} not found")
            return False
        except Exception as e:
            logging.error(f"Error deleting user: {str(e)}")
            raise
    
    def delete_admin(self, admin_id):
        """Delete admin by ID"""
        try:
            wb = openpyxl.load_workbook(self.admins_file)
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == admin_id:
                    ws.delete_rows(row_num)
                    wb.save(self.admins_file)
                    logging.info(f"Deleted admin ID: {admin_id}")
                    return True
            
            logging.warning(f"Admin ID {admin_id} not found")
            return False
        except Exception as e:
            logging.error(f"Error deleting admin: {str(e)}")
            raise
